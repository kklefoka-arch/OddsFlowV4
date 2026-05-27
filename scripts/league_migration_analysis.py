"""League Migration Analysis — V3 policy vs active leagues.

Compares V3 calibration hit rates (European-heavy 27k corpus) against
realised hit rates from the currently active subscription leagues
(non-European / fringe). Flags material variance and writes:

  1. Output/LEAGUE_MIGRATION_ANALYSIS_{date}.xlsx  — full Excel report
  2. Patch to app/engine/static_policy.py          — V3_MIGRATION dict added
     (additive only — V3 policy is never modified)

Usage:
    python scripts/league_migration_analysis.py

Run AFTER scripts/fetch_historical.py has completed.
"""

from __future__ import annotations

import sqlite3
import json
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("WARNING: openpyxl not installed — Excel output skipped. pip install openpyxl")

DB      = r"C:\OddsFlowV4\data\oddsflow_v4.db"
OUT_DIR = Path(r"C:\OddsFlow AI Website\Output")
TODAY   = datetime.now(timezone.utc).strftime("%Y-%m-%d")

# ---------------------------------------------------------------------------
# V3 baseline — directly from static_policy.V3_MARKETS
# ---------------------------------------------------------------------------
V3_BASELINE: dict[tuple[str, str], dict[str, dict]] = {
    ("strong",   "slight_over"):  {"goals_nl": {"hit": 72.2, "n": 4997}},
    ("strong",   "slight_under"): {"goals_nl": {"hit": 66.6, "n": 5925}},
    ("standard", "slight_over"):  {"goals_nl": {"hit": 78.2, "n": 9449}, "corners_nl": {"hit": 64.5, "n": 7274}},
    ("standard", "strong_over"):  {"goals_nl": {"hit": 83.7, "n": 1319}, "corners_nl": {"hit": 69.9, "n": 1173}},
    ("standard", "slight_under"): {"goals_nl": {"hit": 71.6, "n": 1940}, "corners_nl": {"hit": 57.8, "n": 1316}},
    ("low",      "slight_over"):  {"dnb":      {"hit": 84.9, "n": 1733}},
    ("low",      "slight_under"): {"dnb":      {"hit": 91.6, "n":  675}},
    ("one_sided","slight_over"):  {"alpha_win":{"hit": 76.6, "n": 1119}},
    ("one_sided","slight_under"): {"alpha_win":{"hit": 81.0, "n":  814}},
}

# Active league sportmonks IDs with tier
ACTIVE_LEAGUES: dict[int, int] = {
    573:1, 444:1, 345:1, 292:1, 360:1, 779:1, 648:1, 3537:1, 1034:1,
    393:2, 405:2, 579:2, 585:2, 588:2, 681:2, 678:2, 696:2, 1689:2, 295:2, 286:2, 289:2, 791:2, 3550:2, 989:2,
    1642:3, 351:3, 797:3, 1607:3, 2545:3, 1098:3,
}

# Removed (European) leagues — for context
REMOVED_LEAGUES = {
    8:  "England — Premier League",
    301: "France — Ligue 1",
    384: "Italy — Serie A",
    564: "Spain — La Liga",
    567: "Spain — La Liga 2",
}

# Variance thresholds
WATCH_PP    = 3.0   # ≥3pp delta → WATCH
MATERIAL_PP = 6.0   # ≥6pp delta → MATERIAL

# Minimum sample for meaningful comparison
MIN_N = 30

# ---------------------------------------------------------------------------
# Classification helpers (mirrors fetch_upcoming / classify.py)
# ---------------------------------------------------------------------------

def zone_of(d: float | None) -> str | None:
    if d is None: return None
    if d < 2.70:  return None        # excluded
    if d < 3.40:  return "strong"
    if d < 4.10:  return "standard"
    if d < 4.80:  return "low"
    return "one_sided"


def bts_of(y: float | None, n: float | None) -> str | None:
    if y is None or n is None: return None
    if y <= n:
        return "strong_over" if y < 1.50 else "slight_over"
    return "strong_under" if n < 1.50 else "slight_under"


# ---------------------------------------------------------------------------
# Result evaluators
# ---------------------------------------------------------------------------

def _alpha_home(home_odd, away_odd) -> bool:
    if home_odd is None or away_odd is None: return True
    return home_odd <= away_odd


def goals_hit(row: dict, line: float) -> bool | None:
    tg = row.get("total_goals")
    if tg is None:
        hs, aws = row.get("home_score"), row.get("away_score")
        if hs is None or aws is None: return None
        tg = hs + aws
    return tg > line


def corners_hit(row: dict, line: float) -> bool | None:
    tc = row.get("total_corners")
    if tc is None: return None
    return tc > line


def dnb_hit(row: dict) -> bool | None:
    hs, aws = row.get("home_score"), row.get("away_score")
    ho, ao  = row.get("home_odd"),   row.get("away_odd")
    if hs is None or aws is None: return None
    alpha_home = _alpha_home(ho, ao)
    alpha_wins = (hs > aws) if alpha_home else (aws > hs)
    return alpha_wins or (hs == aws)


def alpha_win_hit(row: dict) -> bool | None:
    hs, aws = row.get("home_score"), row.get("away_score")
    ho, ao  = row.get("home_odd"),   row.get("away_odd")
    if hs is None or aws is None: return None
    alpha_home = _alpha_home(ho, ao)
    return (hs > aws) if alpha_home else (aws > hs)


# ---------------------------------------------------------------------------
# Core analysis
# ---------------------------------------------------------------------------

def analyse(conn: sqlite3.Connection) -> dict:
    """Fetch all settled active-league fixtures and compute per-cell hit rates."""

    sm_ids = list(ACTIVE_LEAGUES.keys())
    placeholders = ",".join("?" * len(sm_ids))

    rows = conn.execute(f"""
        SELECT
            f.draw_odd, f.btts_yes_odd, f.btts_no_odd,
            f.home_odd, f.away_odd,
            f.home_score, f.away_score, f.total_goals,
            fs.total_corners,
            f.tier,
            l.sportmonks_id AS league_sm,
            l.name AS league_name, l.country
        FROM fixtures f
        JOIN leagues l ON f.league_id = l.id
        LEFT JOIN fixture_stats fs ON fs.fixture_id = f.id
        WHERE l.sportmonks_id IN ({placeholders})
          AND f.home_score IS NOT NULL
    """, sm_ids).fetchall()

    print(f"Settled active-league fixtures: {len(rows)}")

    # Accumulate per-cell stats
    # Structure: acc[(zone, bts)][market] = {hits, n, corners_n, corners_hits}
    acc: dict[tuple[str,str], dict[str, dict]] = {}

    league_coverage: dict[tuple[str,str], dict[str, int]] = {}  # (country, name) → cell counts

    no_zone = no_bts = 0
    for r in rows:
        d = dict(zip(
            ["draw_odd","btts_yes_odd","btts_no_odd","home_odd","away_odd",
             "home_score","away_score","total_goals","total_corners",
             "tier","league_sm","league_name","country"],
            r
        ))
        zone = zone_of(d["draw_odd"])
        bts  = bts_of(d["btts_yes_odd"], d["btts_no_odd"])
        if zone is None: no_zone += 1; continue
        if bts  is None: no_bts  += 1; continue

        key = (zone, bts)
        if key not in acc:
            acc[key] = {
                "n":             0,
                "goals_hits":    0, "goals_n":   0,
                "corners_hits":  0, "corners_n": 0,
                "dnb_hits":      0, "dnb_n":     0,
                "aw_hits":       0, "aw_n":      0,
                "by_tier":       {1: {"n":0,"goals_h":0,"goals_n":0,"corners_h":0,"corners_n":0,"dnb_h":0,"dnb_n":0,"aw_h":0,"aw_n":0},
                                  2: {"n":0,"goals_h":0,"goals_n":0,"corners_h":0,"corners_n":0,"dnb_h":0,"dnb_n":0,"aw_h":0,"aw_n":0},
                                  3: {"n":0,"goals_h":0,"goals_n":0,"corners_h":0,"corners_n":0,"dnb_h":0,"dnb_n":0,"aw_h":0,"aw_n":0}},
            }
        cell = acc[key]
        tier = d.get("tier") or 1
        tc   = cell["by_tier"].get(tier, cell["by_tier"][1])

        cell["n"] += 1
        tc["n"]   += 1

        # Goals Over 1.5
        g = goals_hit(d, 1.5)
        if g is not None:
            cell["goals_n"] += 1; cell["goals_hits"] += int(g)
            tc["goals_n"]   += 1; tc["goals_h"]      += int(g)

        # Corners Over 8.5
        c = corners_hit(d, 8.5)
        if c is not None:
            cell["corners_n"] += 1; cell["corners_hits"] += int(c)
            tc["corners_n"]   += 1; tc["corners_h"]      += int(c)

        # DNB
        dnb = dnb_hit(d)
        if dnb is not None:
            cell["dnb_n"] += 1; cell["dnb_hits"] += int(dnb)
            tc["dnb_n"]   += 1; tc["dnb_h"]      += int(dnb)

        # Alpha win
        aw = alpha_win_hit(d)
        if aw is not None:
            cell["aw_n"] += 1; cell["aw_hits"] += int(aw)
            tc["aw_n"]   += 1; tc["aw_h"]      += int(aw)

    print(f"  No zone (draw<2.70 excluded): {no_zone}")
    print(f"  No BTS odds:                  {no_bts}")

    def pct(h, n): return round(h / n * 100, 1) if n >= 1 else None
    def delta(a, b): return round(a - b, 1) if (a is not None and b is not None) else None
    def flag(d):
        if d is None: return "NO_DATA"
        ad = abs(d)
        if ad >= MATERIAL_PP: return "MATERIAL"
        if ad >= WATCH_PP:    return "WATCH"
        return "OK"

    results = {}

    for (zone, bts), cell in acc.items():
        baseline_cell = V3_BASELINE.get((zone, bts), {})
        results[(zone, bts)] = {
            "zone": zone, "bts": bts,
            "n_active": cell["n"],
            "markets": {},
        }
        mkt_map = {
            "goals_nl":   (pct(cell["goals_hits"],   cell["goals_n"]),   cell["goals_n"]),
            "corners_nl": (pct(cell["corners_hits"],  cell["corners_n"]), cell["corners_n"]),
            "dnb":        (pct(cell["dnb_hits"],      cell["dnb_n"]),     cell["dnb_n"]),
            "alpha_win":  (pct(cell["aw_hits"],       cell["aw_n"]),      cell["aw_n"]),
        }
        for mkt, (hit_active, n_active) in mkt_map.items():
            bl = baseline_cell.get(mkt, {})
            hit_v3 = bl.get("hit")
            n_v3   = bl.get("n")
            d_pp   = delta(hit_active, hit_v3) if n_active >= MIN_N else None
            results[(zone, bts)]["markets"][mkt] = {
                "hit_active":  hit_active,
                "n_active":    n_active,
                "hit_v3":      hit_v3,
                "n_v3":        n_v3,
                "delta_pp":    d_pp,
                "flag":        flag(d_pp) if n_active >= MIN_N else "LOW_N",
                "in_v3_policy": mkt in baseline_cell,
            }
        # Tier breakdown
        results[(zone, bts)]["by_tier"] = {}
        for tier in (1, 2, 3):
            tc = cell["by_tier"].get(tier, {})
            results[(zone, bts)]["by_tier"][tier] = {
                "n": tc.get("n", 0),
                "goals_hit":   pct(tc.get("goals_h",0),   tc.get("goals_n",0)),
                "corners_hit": pct(tc.get("corners_h",0), tc.get("corners_n",0)),
                "dnb_hit":     pct(tc.get("dnb_h",0),     tc.get("dnb_n",0)),
                "aw_hit":      pct(tc.get("aw_h",0),       tc.get("aw_n",0)),
                "goals_n":     tc.get("goals_n",0),
                "corners_n":   tc.get("corners_n",0),
                "dnb_n":       tc.get("dnb_n",0),
                "aw_n":        tc.get("aw_n",0),
            }

    return results


def league_breakdown(conn: sqlite3.Connection) -> list[dict]:
    """Per-league fixture counts and basic hit rates."""
    sm_ids = list(ACTIVE_LEAGUES.keys())
    rows = conn.execute(f"""
        SELECT l.country, l.name, l.tier,
               COUNT(f.id) as n,
               ROUND(100.0*SUM(CASE WHEN f.total_goals > 1.5 THEN 1 ELSE 0 END)/COUNT(f.id),1) as g15_pct,
               ROUND(100.0*SUM(CASE WHEN f.draw_zone IS NOT NULL THEN 1 ELSE 0 END)/COUNT(f.id),1) as has_zone_pct,
               ROUND(100.0*SUM(CASE WHEN f.bts_pocket IS NOT NULL THEN 1 ELSE 0 END)/COUNT(f.id),1) as has_bts_pct,
               ROUND(100.0*SUM(CASE WHEN fs.total_corners IS NOT NULL THEN 1 ELSE 0 END)/COUNT(f.id),1) as has_corners_pct
        FROM fixtures f
        JOIN leagues l ON f.league_id = l.id
        LEFT JOIN fixture_stats fs ON fs.fixture_id = f.id
        WHERE l.sportmonks_id IN ({','.join('?'*len(sm_ids))})
          AND f.home_score IS NOT NULL
        GROUP BY l.id
        ORDER BY l.tier, n DESC
    """, sm_ids).fetchall()
    return [dict(zip(["country","name","tier","n","g15_pct","has_zone_pct","has_bts_pct","has_corners_pct"], r)) for r in rows]


def calibration_corpus_summary(conn: sqlite3.Connection) -> dict:
    """Summarise the European calibration corpus for context."""
    removed_sm = list(REMOVED_LEAGUES.keys())
    total = conn.execute("SELECT COUNT(*) FROM fixtures WHERE home_score IS NOT NULL").fetchone()[0]
    active_sm = list(ACTIVE_LEAGUES.keys())
    active_n = conn.execute(f"""
        SELECT COUNT(*) FROM fixtures f JOIN leagues l ON f.league_id=l.id
        WHERE l.sportmonks_id IN ({','.join('?'*len(active_sm))}) AND f.home_score IS NOT NULL
    """, active_sm).fetchone()[0]
    euro_n = conn.execute(f"""
        SELECT COUNT(*) FROM fixtures f JOIN leagues l ON f.league_id=l.id
        WHERE l.sportmonks_id IN ({','.join('?'*len(removed_sm))}) AND f.home_score IS NOT NULL
    """, removed_sm).fetchone()[0]
    no_sm = total - active_n - euro_n
    return {
        "total_settled": total,
        "active_leagues_n": active_n,
        "removed_euro_leagues_n": euro_n,
        "other_seeded_n": no_sm,
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

ZONE_ORDER = ["strong","standard","low","one_sided"]
BTS_ORDER  = ["slight_over","strong_over","slight_under","strong_under"]

FILL_MATERIAL = PatternFill("solid", fgColor="FF6B6B") if HAS_EXCEL else None
FILL_WATCH    = PatternFill("solid", fgColor="FFD93D") if HAS_EXCEL else None
FILL_OK       = PatternFill("solid", fgColor="6BCB77") if HAS_EXCEL else None
FILL_NODATA   = PatternFill("solid", fgColor="CCCCCC") if HAS_EXCEL else None
FILL_HDRBLUE  = PatternFill("solid", fgColor="264653") if HAS_EXCEL else None
FILL_HDRGRAY  = PatternFill("solid", fgColor="457B9D") if HAS_EXCEL else None
BOLD          = Font(bold=True) if HAS_EXCEL else None
WHITE         = Font(bold=True, color="FFFFFF") if HAS_EXCEL else None

def _hdr(ws, row, col, val, hdr=True):
    c = ws.cell(row=row, column=col, value=val)
    if hdr:
        c.fill  = FILL_HDRBLUE
        c.font  = WHITE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    return c

def _flag_fill(flag):
    if flag == "MATERIAL": return FILL_MATERIAL
    if flag == "WATCH":    return FILL_WATCH
    if flag == "OK":       return FILL_OK
    return FILL_NODATA


def write_excel(results: dict, league_data: list, corpus_summary: dict, path: Path):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Executive Summary ----
    ws = wb.active
    ws.title = "00_Summary"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.cell(r, 1, "LEAGUE MIGRATION ANALYSIS").font = Font(bold=True, size=14)
    r += 1
    ws.cell(r, 1, f"Generated: {TODAY}")
    r += 2

    ws.cell(r, 1, "CORPUS BREAKDOWN").font = BOLD; r += 1
    ws.cell(r, 1, "Total settled fixtures in DB:");         ws.cell(r, 2, corpus_summary["total_settled"]); r += 1
    ws.cell(r, 1, "Active subscription leagues (settled):"); ws.cell(r, 2, corpus_summary["active_leagues_n"]); r += 1
    ws.cell(r, 1, "Removed EU leagues (still seeded):"); ws.cell(r, 2, corpus_summary["removed_euro_leagues_n"]); r += 1
    ws.cell(r, 1, "Other seeded fixtures (no SM ID):"); ws.cell(r, 2, corpus_summary["other_seeded_n"]); r += 2

    ws.cell(r, 1, "MIGRATION CONTEXT").font = BOLD; r += 1
    ws.cell(r, 1, "V3 calibrated on:"); ws.cell(r, 2, "European-heavy corpus (PL, Serie A, Ligue 1, La Liga)"); r += 1
    ws.cell(r, 1, "Active leagues are:"); ws.cell(r, 2, "Americas, Asia, Scandinavia, Baltic — NO big-5 EU"); r += 1
    ws.cell(r, 1, "Removed (budget cut):"); ws.cell(r, 2, "PL, Ligue 1, Serie A, La Liga, La Liga 2"); r += 1
    ws.cell(r, 1, "Added (replacements):"); ws.cell(r, 2, "Argentina Reserves, Iceland, Kazakhstan, Lithuania, USL League Two"); r += 2

    ws.cell(r, 1, "VARIANCE FLAGS").font = BOLD; r += 1
    ws.cell(r, 1, "OK (green):"); ws.cell(r, 2, "delta < 3pp — policy holds"); r += 1
    ws.cell(r, 1, "WATCH (yellow):"); ws.cell(r, 2, "3-6pp — monitor closely"); r += 1
    ws.cell(r, 1, "MATERIAL (red):"); ws.cell(r, 2, f">={MATERIAL_PP}pp — significant divergence from V3 baseline"); r += 1
    ws.cell(r, 1, "LOW_N (grey):"); ws.cell(r, 2, f"n<{MIN_N} — insufficient data for comparison"); r += 2

    # Count flagged cells
    material_cells = []
    watch_cells    = []
    ok_cells       = []
    for (zone, bts), res in results.items():
        for mkt, mdata in res["markets"].items():
            if not mdata["in_v3_policy"]: continue
            f = mdata["flag"]
            label = f"{zone}:{bts} {mkt}"
            if f == "MATERIAL": material_cells.append((label, mdata["delta_pp"]))
            elif f == "WATCH":  watch_cells.append((label, mdata["delta_pp"]))
            elif f == "OK":     ok_cells.append(label)

    ws.cell(r, 1, "VERDICT").font = Font(bold=True, size=12)
    if not material_cells:
        ws.cell(r, 2, "NO MATERIAL VARIANCE — V3 policy holds for active leagues").font = Font(color="006400", bold=True)
    else:
        ws.cell(r, 2, f"{len(material_cells)} MATERIAL cell(s) — see migration layer").font = Font(color="CC0000", bold=True)
    r += 2

    ws.cell(r, 1, "MATERIAL cells (policy-relevant picks):").font = BOLD; r += 1
    for label, dp in material_cells:
        ws.cell(r, 1, label); ws.cell(r, 2, f"{dp:+.1f}pp"); r += 1
    if not material_cells:
        ws.cell(r, 1, "None"); r += 1
    r += 1
    ws.cell(r, 1, "WATCH cells:").font = BOLD; r += 1
    for label, dp in watch_cells:
        ws.cell(r, 1, label); ws.cell(r, 2, f"{dp:+.1f}pp"); r += 1
    if not watch_cells:
        ws.cell(r, 1, "None"); r += 1

    # ---- Sheet 2: Cell Comparison Matrix ----
    ws2 = wb.create_sheet("01_Cell_Comparison")
    MKTS_ORDERED = ["goals_nl","corners_nl","dnb","alpha_win"]
    headers = ["Zone","BTS Pocket","N (Active)","Market",
               "V3 Hit%","V3 N","Active Hit%","Active N","Delta pp","Flag","In V3 Policy"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws2, 1, ci, h)
    ws2.row_dimensions[1].height = 30

    row = 2
    for zone in ZONE_ORDER:
        for bts in BTS_ORDER:
            res = results.get((zone, bts))
            if res is None: continue
            for mkt in MKTS_ORDERED:
                md = res["markets"].get(mkt, {})
                fl = md.get("flag","—")
                cells_data = [
                    zone, bts, res["n_active"], mkt,
                    md.get("hit_v3","—"), md.get("n_v3","—"),
                    md.get("hit_active","—"), md.get("n_active","—"),
                    md.get("delta_pp","—"), fl,
                    "YES" if md.get("in_v3_policy") else "no",
                ]
                for ci, val in enumerate(cells_data, 1):
                    c = ws2.cell(row=row, column=ci, value=val)
                    if ci == 10:   # Flag column
                        c.fill = _flag_fill(fl)
                        if fl in ("MATERIAL","WATCH","OK"): c.font = Font(bold=True)
                row += 1

    for col in range(1, len(headers)+1):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    ws2.freeze_panes = "A2"

    # ---- Sheet 3: Tier Breakdown ----
    ws3 = wb.create_sheet("02_Tier_Breakdown")
    t_headers = ["Zone","BTS","Tier","N","Goals Hit%","Goals N","Corners Hit%","Corners N","DNB Hit%","DNB N","AW Hit%","AW N"]
    for ci, h in enumerate(t_headers, 1):
        _hdr(ws3, 1, ci, h)
    row = 2
    for zone in ZONE_ORDER:
        for bts in BTS_ORDER:
            res = results.get((zone,bts))
            if res is None: continue
            for tier in (1,2,3):
                td = res["by_tier"].get(tier, {})
                row_data = [zone, bts, f"T{tier}", td.get("n",0),
                            td.get("goals_hit","—"), td.get("goals_n",0),
                            td.get("corners_hit","—"), td.get("corners_n",0),
                            td.get("dnb_hit","—"), td.get("dnb_n",0),
                            td.get("aw_hit","—"), td.get("aw_n",0)]
                for ci, val in enumerate(row_data, 1):
                    ws3.cell(row=row, column=ci, value=val)
                row += 1
    for col in range(1, len(t_headers)+1):
        ws3.column_dimensions[get_column_letter(col)].width = 14
    ws3.freeze_panes = "A2"

    # ---- Sheet 4: League Breakdown ----
    ws4 = wb.create_sheet("03_League_Breakdown")
    l_headers = ["Country","League","Tier","N Settled","Goals>1.5%","Has Zone%","Has BTS%","Has Corners%"]
    for ci, h in enumerate(l_headers, 1):
        _hdr(ws4, 1, ci, h)
    for ri, ld in enumerate(league_data, 2):
        row_data = [ld["country"], ld["name"], f"T{ld['tier']}", ld["n"],
                    ld["g15_pct"], ld["has_zone_pct"], ld["has_bts_pct"], ld["has_corners_pct"]]
        for ci, val in enumerate(row_data, 1):
            ws4.cell(row=ri, column=ci, value=val)
    for col in range(1, len(l_headers)+1):
        ws4.column_dimensions[get_column_letter(col)].width = 20
    ws4.freeze_panes = "A2"

    # ---- Sheet 5: Migration Policy (what goes in static_policy.py) ----
    ws5 = wb.create_sheet("04_Migration_Layer")
    ws5.cell(1, 1, "V3_MIGRATION — additive dict for static_policy.py").font = Font(bold=True, size=12)
    ws5.cell(2, 1, "This dict is INFORMATIONAL. V3_ACTIVE is unchanged.")
    ws5.cell(3, 1, "Cells with MATERIAL or WATCH flag are listed with their active-league rates.")
    ws5.cell(4, 1, "Engine will use active_hit% for monitoring/drift detection on these leagues.")
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 20
    r = 6
    for zone in ZONE_ORDER:
        for bts in BTS_ORDER:
            res = results.get((zone,bts))
            if res is None: continue
            for mkt, md in res["markets"].items():
                if not md.get("in_v3_policy"): continue
                if md.get("flag") in ("MATERIAL","WATCH","OK","LOW_N"):
                    ws5.cell(r, 1, f"({zone!r}, {bts!r}) {mkt}")
                    ws5.cell(r, 2, f"active={md.get('hit_active','?')}%  v3={md.get('hit_v3','?')}%  delta={md.get('delta_pp','?')}pp  flag={md.get('flag','?')}")
                    c = ws5.cell(r, 2)
                    fl = md.get("flag","")
                    if fl == "MATERIAL": c.fill = FILL_MATERIAL
                    elif fl == "WATCH":  c.fill = FILL_WATCH
                    elif fl == "OK":     c.fill = FILL_OK
                    r += 1

    wb.save(path)
    print(f"Excel saved: {path}")


# ---------------------------------------------------------------------------
# Generate V3_MIGRATION Python code
# ---------------------------------------------------------------------------

def generate_migration_code(results: dict) -> str:
    lines = [
        "",
        "# ---------------------------------------------------------------------------",
        "# V3_MIGRATION — additive monitoring layer for active (non-European) leagues.",
        "# Added: " + TODAY,
        "#",
        "# Context: V3 policy calibrated on European-heavy 27k corpus",
        "#   (PL, Serie A, Ligue 1, La Liga). Active subscription migrated to",
        "#   Americas / Asia / Scandinavia / Baltic (budget reallocation, 2026-05).",
        "#",
        "# V3_ACTIVE and V3_MARKETS are UNCHANGED. V3_MIGRATION records realised",
        "# hit rates for the current league set so drift monitoring can compare",
        "# active performance against the correct baseline per league group.",
        "#",
        "# When/if European leagues return, delete V3_MIGRATION and revert to",
        "# using V3_MARKETS hit rates as the sole monitoring baseline.",
        "# ---------------------------------------------------------------------------",
        "",
        "V3_MIGRATION: dict[tuple[str, str], dict[str, dict]] = {",
    ]

    for zone in ZONE_ORDER:
        for bts in BTS_ORDER:
            res = results.get((zone, bts))
            if res is None:
                continue
            has_policy_mkt = any(md.get("in_v3_policy") for md in res["markets"].values())
            if not has_policy_mkt:
                continue
            lines.append(f'    ("{zone}", "{bts}"): {{')
            for mkt, md in res["markets"].items():
                if not md.get("in_v3_policy"):
                    continue
                ha  = md.get("hit_active", "None")
                na  = md.get("n_active", 0)
                hv  = md.get("hit_v3", "None")
                dp  = md.get("delta_pp", "None")
                fl  = md.get("flag", "LOW_N")
                lines.append(f'        "{mkt}": {{')
                lines.append(f'            "hit_active_leagues": {ha},  # realised on current subscription')
                lines.append(f'            "n_active": {na},')
                lines.append(f'            "hit_v3_baseline": {hv},  # from V3_MARKETS (European corpus)')
                lines.append(f'            "delta_pp": {dp},')
                lines.append(f'            "flag": "{fl}",  # OK / WATCH / MATERIAL / LOW_N')
                lines.append(f'        }},')
            lines.append("    },")

    lines.append("}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Patch static_policy.py
# ---------------------------------------------------------------------------

def patch_static_policy(migration_code: str):
    policy_path = Path(r"C:\OddsFlowV4\app\engine\static_policy.py")
    content = policy_path.read_text(encoding="utf-8")

    marker = "# V3_MIGRATION"
    if marker in content:
        # Already patched — replace existing block
        idx = content.index(marker)
        # Find the start of the block (look back for the dashes line)
        dash_idx = content.rfind("# -----", 0, idx)
        if dash_idx > 0:
            content = content[:dash_idx].rstrip() + "\n" + migration_code + "\n"
        else:
            content = content[:idx].rstrip() + "\n" + migration_code + "\n"
        print("  Replaced existing V3_MIGRATION block in static_policy.py")
    else:
        # Append to end
        content = content.rstrip() + "\n" + migration_code + "\n"
        print("  Appended V3_MIGRATION to static_policy.py")

    policy_path.write_text(content, encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row

    print("=" * 60)
    print("LEAGUE MIGRATION ANALYSIS")
    print(f"Date: {TODAY}")
    print("=" * 60)

    corpus = calibration_corpus_summary(conn)
    print(f"\nCorpus: {corpus['total_settled']:,} settled total")
    print(f"  Active leagues:      {corpus['active_leagues_n']:,}")
    print(f"  Removed EU leagues:  {corpus['removed_euro_leagues_n']:,}")
    print(f"  Other seeded:        {corpus['other_seeded_n']:,}")

    print("\nRunning per-cell analysis on active leagues...")
    results = analyse(conn)
    league_data = league_breakdown(conn)
    conn.close()

    # Print console summary
    print(f"\n{'='*60}")
    print(f"CELL COMPARISON — V3 Baseline vs Active Leagues")
    print(f"{'='*60}")
    print(f"{'Zone:BTS':<28} {'Market':<12} {'V3%':>6} {'Act%':>6} {'Delta':>7} {'Flag':<10}")
    print("-"*75)

    material_count = watch_count = ok_count = low_n_count = 0
    for zone in ZONE_ORDER:
        for bts in BTS_ORDER:
            res = results.get((zone, bts))
            if res is None: continue
            for mkt, md in res["markets"].items():
                if not md.get("in_v3_policy"): continue
                v3h  = f"{md['hit_v3']:.1f}" if md.get("hit_v3") else "—"
                ach  = f"{md['hit_active']:.1f}" if md.get("hit_active") else "—"
                dp   = f"{md['delta_pp']:+.1f}" if md.get("delta_pp") is not None else "—"
                fl   = md.get("flag","—")
                label = f"{zone}:{bts}"
                print(f"  {label:<26} {mkt:<12} {v3h:>6} {ach:>6} {dp:>7}  {fl}")
                if fl == "MATERIAL": material_count += 1
                elif fl == "WATCH":  watch_count += 1
                elif fl == "OK":     ok_count += 1
                else:                low_n_count += 1

    print(f"\nSummary: {ok_count} OK  |  {watch_count} WATCH  |  {material_count} MATERIAL  |  {low_n_count} LOW_N")

    # Generate migration code
    migration_code = generate_migration_code(results)
    print("\nGenerating V3_MIGRATION dict and patching static_policy.py...")
    patch_static_policy(migration_code)

    # Excel output
    if HAS_EXCEL:
        out_path = OUT_DIR / f"LEAGUE_MIGRATION_ANALYSIS_{TODAY}.xlsx"
        write_excel(results, league_data, corpus, out_path)
    else:
        print("Skipping Excel — openpyxl not available")

    # Save JSON for reference
    json_out = OUT_DIR / f"LEAGUE_MIGRATION_ANALYSIS_{TODAY}.json"
    serialisable = {f"{z}:{b}": v for (z,b),v in results.items()}
    json_out.write_text(json.dumps(serialisable, indent=2), encoding="utf-8")
    print(f"JSON saved: {json_out}")

    print("\nDone. V3_MIGRATION added to static_policy.py (additive — V3_ACTIVE unchanged).")
    print("Restart server to reload: uvicorn app.main:app --host 127.0.0.1 --port 8000")


if __name__ == "__main__":
    main()
